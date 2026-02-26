# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1,<4", "click>=8.3,<9"]
# ///
"""Excel (.xlsx) operations tool.

Commands: read, write, info, add-sheet, export.

Usage:
    uv run xlsx_tool.py COMMAND [OPTIONS]
"""

import csv
import io
import json
import sys
from pathlib import Path

import click
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string


def write_output(text: str, output_path: str | None) -> None:
    """Write text to file or stdout."""
    if output_path:
        Path(output_path).write_text(text, encoding="utf-8")
        click.echo(f"Output written to {output_path}", err=True)
    else:
        click.echo(text)


def parse_cell_ref(ref: str) -> tuple[int, int]:
    """Parse a cell reference like 'A1' into (row, col) 1-based."""
    col_str = ""
    row_str = ""
    for ch in ref:
        if ch.isalpha():
            col_str += ch
        else:
            row_str += ch
    col = column_index_from_string(col_str.upper())
    row = int(row_str)
    return row, col


@click.group()
def cli() -> None:
    """Excel (.xlsx) operations tool."""
    pass


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--range", "cell_range", type=str, default=None, help="Cell range, e.g. 'A1:C10'.")
@click.option("--format", "fmt", type=click.Choice(["text", "csv", "json"]), default="text", help="Output format.")
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def read(file: str, sheet: str | None, cell_range: str | None, fmt: str, output: str | None) -> None:
    """Read cells, ranges, or entire sheets from an Excel file."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        if sheet:
            if sheet not in wb.sheetnames:
                click.echo(f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}", err=True)
                sys.exit(1)
            ws = wb[sheet]
        else:
            ws = wb.active

        if cell_range:
            rows = list(ws[cell_range])
        else:
            rows = list(ws.iter_rows())

        if not rows:
            write_output("(empty)", output)
            wb.close()
            return

        data: list[list[object]] = []
        for row in rows:
            data.append([cell.value for cell in row])

        if fmt == "json":
            # Use first row as headers if possible
            if len(data) > 1:
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
                records = []
                for row_data in data[1:]:
                    record: dict[str, object] = {}
                    for i, val in enumerate(row_data):
                        key = headers[i] if i < len(headers) else f"col_{i}"
                        record[key] = val
                    records.append(record)
                result = json.dumps(records, indent=2, default=str)
            else:
                result = json.dumps(data, indent=2, default=str)
        elif fmt == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            for row_data in data:
                writer.writerow(row_data)
            result = buf.getvalue()
        else:
            lines: list[str] = []
            # Calculate column widths for aligned text output
            str_data = [[str(v) if v is not None else "" for v in row] for row in data]
            if str_data:
                max_cols = max(len(row) for row in str_data)
                col_widths = [0] * max_cols
                for row in str_data:
                    for i, val in enumerate(row):
                        col_widths[i] = max(col_widths[i], len(val))
                for row in str_data:
                    parts = []
                    for i, val in enumerate(row):
                        width = col_widths[i] if i < len(col_widths) else 0
                        parts.append(val.ljust(width))
                    lines.append("  ".join(parts).rstrip())
            result = "\n".join(lines)

        wb.close()
        write_output(result, output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--cell", type=str, required=True, help="Cell reference, e.g. 'A1'.")
@click.option("--value", type=str, required=True, help="Value to write.")
@click.option("--type", "val_type", type=click.Choice(["string", "number", "bool"]), default="string", help="Value type.")
def write(file: str, sheet: str | None, cell: str, value: str, val_type: str) -> None:
    """Write a value to a specific cell in an Excel file.

    Creates the file if it does not exist.
    """
    try:
        file_path = Path(file)
        if file_path.exists():
            wb = load_workbook(file)
        else:
            wb = Workbook()

        if sheet:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
            ws = wb[sheet]
        else:
            ws = wb.active

        # Convert value type
        converted: object
        if val_type == "number":
            try:
                converted = int(value)
            except ValueError:
                converted = float(value)
        elif val_type == "bool":
            converted = value.lower() in ("true", "1", "yes")
        else:
            converted = value

        ws[cell.upper()] = converted
        wb.save(file)
        wb.close()
        click.echo(f"Wrote '{converted}' to {cell.upper()} in {file_path.name}", err=True)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def info(file: str, output: str | None) -> None:
    """Show workbook information: sheets, dimensions, cell counts."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets_info.append({
                "name": name,
                "dimensions": ws.dimensions,
                "min_row": ws.min_row,
                "max_row": ws.max_row,
                "min_column": ws.min_column,
                "max_column": ws.max_column,
            })

        info_dict = {
            "file": str(Path(file).resolve()),
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_info,
        }

        wb.close()
        write_output(json.dumps(info_dict, indent=2, default=str), output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command("add-sheet")
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--name", type=str, required=True, help="Name for the new sheet.")
@click.option("--position", type=int, default=None, help="Position index (0-based). Default: append at end.")
def add_sheet(file: str, name: str, position: int | None) -> None:
    """Add a new sheet to an Excel file.

    Creates the file if it does not exist.
    """
    try:
        file_path = Path(file)
        if file_path.exists():
            wb = load_workbook(file)
        else:
            wb = Workbook()
            # Remove default sheet if creating new file
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if name in wb.sheetnames:
            click.echo(f"Error: sheet '{name}' already exists.", err=True)
            sys.exit(1)

        if position is not None:
            wb.create_sheet(name, position)
        else:
            wb.create_sheet(name)

        wb.save(file)
        wb.close()
        click.echo(f"Added sheet '{name}' to {file_path.name}", err=True)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--format", "fmt", type=click.Choice(["csv", "json"]), default="csv", help="Export format.")
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def export(file: str, sheet: str | None, fmt: str, output: str | None) -> None:
    """Export a sheet as CSV or JSON."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        if sheet:
            if sheet not in wb.sheetnames:
                click.echo(f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}", err=True)
                sys.exit(1)
            ws = wb[sheet]
        else:
            ws = wb.active

        rows = list(ws.iter_rows())
        if not rows:
            write_output("(empty)", output)
            wb.close()
            return

        data = [[cell.value for cell in row] for row in rows]

        if fmt == "json":
            if len(data) > 1:
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
                records = []
                for row_data in data[1:]:
                    record: dict[str, object] = {}
                    for i, val in enumerate(row_data):
                        key = headers[i] if i < len(headers) else f"col_{i}"
                        record[key] = val
                    records.append(record)
                result = json.dumps(records, indent=2, default=str)
            else:
                result = json.dumps(data, indent=2, default=str)
        else:
            buf = io.StringIO()
            writer = csv.writer(buf)
            for row_data in data:
                writer.writerow(row_data)
            result = buf.getvalue()

        wb.close()
        write_output(result, output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


if __name__ == "__main__":
    cli()
