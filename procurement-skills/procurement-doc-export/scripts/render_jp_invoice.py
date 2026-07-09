#!/usr/bin/env python3
"""日本請求書（イノ間請求書 / JPY）渲染：JSON context → blank.xlsx → 可编辑 .xlsx。

独立于美金 PI（render_pi.py）。模板：templates/followup/jp-ino-invoice/blank.xlsx
Sheet「样板」。cell-map 由 blank + samples 实测（见 tests + scratch notes）。

用法：
  uv run --with openpyxl python3 render_jp_invoice.py \\
      --template templates/followup/jp-ino-invoice/blank.xlsx \\
      --out /tmp/jp_out.xlsx \\
      --data ctx.json
  # --data 省略则读 stdin

ctx.json：
  {
    "invoice_no": "OD20260706341",
    "invoice_date": "2026年7月7日",   # 或 2026-07-07 → 自动格式化
    "subject": "電子部品の緊急調達",
    "fx_rmb": 22.1,                   # 任意；缺省保留模板默认
    "fx_usd": 158.2,
    "items": [
      {
        "part": "ANSI/VITA 46.10-2009",
        "qty": 1,
        "unit_price_rmb": null,        # 仕入 S 列
        "unit_price_usd": 50,          # 仕入 T 列
        "unit_price_jpy": null         # 仕入 U 列（直填 JPY）
      }
    ],
    "shipping": {                      # 可选 立替運送料
      "qty": 1,
      "unit_price_rmb": null,
      "unit_price_usd": null,
      "unit_price_jpy": 0
    },
    "import_tax_jpy": 0                # 可选 立替輸入消費税 支付额
  }
成功时把输出路径打到 stdout。
"""
from __future__ import annotations

import argparse
import copy
import json
import os
import re
import sys
from typing import Any

# blank sheet name (measured)
TPL_SHEET = "样板"
PRODUCT_START = 19
PRODUCT_SLOTS = 3  # blank 预留 商品1..3
# blank: rows 22=運送料, 23=FAR手数料, 24=輸入消費税; footer from 27
FEE_COUNT = 3
GAP_AFTER_FEE = 2  # blank 25–26 empty before 振込先@27


def _set(ws, coord: str, value: Any) -> None:
    ws[coord] = value


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        if s.has_style:
            d.font = copy.copy(s.font)
            d.fill = copy.copy(s.fill)
            d.border = copy.copy(s.border)
            d.alignment = copy.copy(s.alignment)
            d.number_format = s.number_format
            d.protection = copy.copy(s.protection)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _insert_rows_keep_merges(ws, idx: int, amount: int) -> None:
    from openpyxl.utils import range_boundaries

    moved = []
    for rng in list(ws.merged_cells.ranges):
        minc, minr, maxc, maxr = range_boundaries(str(rng))
        if minr >= idx:
            ws.unmerge_cells(str(rng))
            moved.append((minc, minr, maxc, maxr))
    ws.insert_rows(idx, amount)
    for minc, minr, maxc, maxr in moved:
        ws.merge_cells(
            start_row=minr + amount,
            start_column=minc,
            end_row=maxr + amount,
            end_column=maxc,
        )


def _delete_rows_keep_merges(ws, idx: int, amount: int) -> None:
    from openpyxl.utils import range_boundaries

    moved = []
    for rng in list(ws.merged_cells.ranges):
        minc, minr, maxc, maxr = range_boundaries(str(rng))
        if minr >= idx + amount:
            ws.unmerge_cells(str(rng))
            moved.append((minc, minr - amount, maxc, maxr - amount))
        elif minr >= idx:
            ws.unmerge_cells(str(rng))
    ws.delete_rows(idx, amount)
    for minc, minr, maxc, maxr in moved:
        ws.merge_cells(start_row=minr, start_column=minc, end_row=maxr, end_column=maxc)


def _merge_product_row(ws, row: int) -> None:
    """Product row merges measured on blank: C:E, F:H, J:L, M:N."""
    for rng in (f"C{row}:E{row}", f"F{row}:H{row}", f"J{row}:L{row}", f"M{row}:N{row}"):
        try:
            ws.merge_cells(rng)
        except Exception:
            pass


def _format_date(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    if "年" in s:
        return s if s.startswith("：") else f"：{s}"
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.group(1), int(m.group(2)), int(m.group(3))
        return f"：{y}年{mo}月{d}日"
    if s.startswith("："):
        return s
    return f"：{s}"


def _format_invoice_no(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    if s.startswith("："):
        return s
    return f"：{s}"


def _write_product_formulas(ws, r: int, rate_row: int) -> None:
    """Per-product formulas measured on blank row 19."""
    _set(ws, f"F{r}", f"=R{r}")
    _set(ws, f"I{r}", "PCS")
    _set(ws, f"J{r}", f"=V{r}")
    _set(ws, f"M{r}", f"=J{r}*F{r}")
    _set(ws, f"V{r}", f"=S{r}*$V${rate_row}+T{r}*$W${rate_row}+U{r}")
    _set(ws, f"W{r}", f"=R{r}*V{r}")


def _write_product_row(ws, r: int, idx: int, it: dict, rate_row: int) -> None:
    _set(ws, f"B{r}", idx)
    _set(ws, f"C{r}", it.get("part") or it.get("name") or "")
    _set(ws, f"Q{r}", f"商品{idx}")
    qty = it.get("qty")
    if qty is None:
        qty = 0
    _set(ws, f"R{r}", qty)
    # clear all price inputs then set provided
    _set(ws, f"S{r}", None)
    _set(ws, f"T{r}", None)
    _set(ws, f"U{r}", None)
    if it.get("unit_price_rmb") is not None:
        _set(ws, f"S{r}", it["unit_price_rmb"])
    if it.get("unit_price_usd") is not None:
        _set(ws, f"T{r}", it["unit_price_usd"])
    if it.get("unit_price_jpy") is not None:
        _set(ws, f"U{r}", it["unit_price_jpy"])
    _write_product_formulas(ws, r, rate_row)


def _adjust_product_slots(ws, n: int, max_col: int) -> None:
    """Make product block length == n (insert before fee block / delete extra slots)."""
    if n < 1:
        raise SystemExit("items 为空：至少要有一个货品行")
    if n > PRODUCT_SLOTS:
        extra = n - PRODUCT_SLOTS
        fee_start = PRODUCT_START + PRODUCT_SLOTS  # 22
        _insert_rows_keep_merges(ws, fee_start, extra)
        for i in range(extra):
            dst = PRODUCT_START + PRODUCT_SLOTS + i
            _copy_row_style(ws, PRODUCT_START, dst, max_col)
            _merge_product_row(ws, dst)
    elif n < PRODUCT_SLOTS:
        # delete unused product rows just above fee block
        _delete_rows_keep_merges(ws, PRODUCT_START + n, PRODUCT_SLOTS - n)


def render(ctx: dict, template_path: str, out_path: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "openpyxl 未安装。请用：uv run --with openpyxl python3 <本脚本> ..."
        )

    if not os.path.exists(template_path):
        raise SystemExit(f"模板不存在：{template_path}")

    items = ctx.get("items") or []
    n = len(items)
    if n < 1:
        raise SystemExit("items 为空：至少要有一个货品行")

    wb = load_workbook(template_path)
    if TPL_SHEET not in wb.sheetnames:
        raise SystemExit(f"模板缺少 sheet「{TPL_SHEET}」（现有：{wb.sheetnames}）")
    ws = wb[TPL_SHEET]
    for name in list(wb.sheetnames):
        if name != TPL_SHEET:
            del wb[name]

    max_col = ws.max_column or 30
    _adjust_product_slots(ws, n, max_col)

    # After adjust: products [19, 18+n], fees start at 19+n
    # blank anchors (n=3): fee@22-24, footer 振込先/小計@27, 税率/売上@28, 合計/仕入@29,
    # 利益@30, 深圳利益@31, 日本利益@32. openpyxl does NOT rewrite formulas on insert/delete.
    fee_ship = PRODUCT_START + n
    fee_far = PRODUCT_START + n + 1
    fee_tax = PRODUCT_START + n + 2
    footer_start = PRODUCT_START + n + FEE_COUNT + GAP_AFTER_FEE  # 振込先 + 小計
    subtotal_row = footer_start
    rate_row = footer_start + 1  # 消費税 + 為替レート V/W + 売上(税抜) R
    tax_row = rate_row
    total_row = footer_start + 2  # 合計 M + 仕入総額 R
    purchase_sum_row = total_row
    profit_row = footer_start + 3  # 利益総額 R
    far_profit_row = footer_start + 4  # イノ深圳利益 R; V = 輸入消費税支払い手数料
    jp_profit_row = footer_start + 5  # イノ日本利益 R; V = 関税等
    last_product = PRODUCT_START + n - 1

    # ── header ──
    _set(ws, "K4", _format_date(ctx.get("invoice_date")))
    _set(ws, "K5", _format_invoice_no(ctx.get("invoice_no")))
    subject = ctx.get("subject") or "電子部品の緊急調達"
    if not str(subject).startswith("件名"):
        subject = f"件名： {subject}"
    _set(ws, "A13", subject)

    # ── products ──
    for i, it in enumerate(items):
        r = PRODUCT_START + i
        _write_product_row(ws, r, i + 1, it, rate_row)

    # ── shipping row (立替運送料) ──
    ship = ctx.get("shipping") or {}
    _set(ws, f"C{fee_ship}", "立替運送料")
    _set(ws, f"Q{fee_ship}", "運送料")
    _set(ws, f"I{fee_ship}", "回")
    sq = ship.get("qty")
    if sq is None:
        sq = 0
    _set(ws, f"R{fee_ship}", sq)
    _set(ws, f"F{fee_ship}", f"=R{fee_ship}")
    _set(ws, f"J{fee_ship}", f"=V{fee_ship}")
    _set(ws, f"M{fee_ship}", f"=J{fee_ship}*F{fee_ship}")
    _set(ws, f"S{fee_ship}", ship.get("unit_price_rmb"))
    _set(ws, f"T{fee_ship}", ship.get("unit_price_usd"))
    _set(ws, f"U{fee_ship}", ship.get("unit_price_jpy"))
    _set(
        ws,
        f"V{fee_ship}",
        f"=S{fee_ship}*$V${rate_row}+T{fee_ship}*$W${rate_row}+U{fee_ship}",
    )

    # ── FAR fee (単価 = イノ深圳利益 R{far_profit_row}) ──
    _set(ws, f"C{fee_far}", "FAR手数料")
    _set(ws, f"F{fee_far}", 1)
    _set(ws, f"I{fee_far}", "一式")
    _set(ws, f"J{fee_far}", f"=$R${far_profit_row}")
    _set(ws, f"M{fee_far}", f"=J{fee_far}*F{fee_far}")
    _set(ws, f"R{fee_far}", "数量")
    _set(ws, f"V{fee_far}", "支払額")

    # ── import tax ──
    _set(ws, f"C{fee_tax}", "立替輸入消費税")
    _set(ws, f"Q{fee_tax}", "輸入消費税")
    _set(ws, f"I{fee_tax}", "回")
    _set(ws, f"R{fee_tax}", 0)
    _set(ws, f"F{fee_tax}", f"=R{fee_tax}")
    tax_jpy = ctx.get("import_tax_jpy")
    if tax_jpy is None:
        tax_jpy = 0
    _set(ws, f"V{fee_tax}", tax_jpy)
    _set(ws, f"J{fee_tax}", f"=$V${fee_tax}")
    _set(ws, f"M{fee_tax}", f"=J{fee_tax}*F{fee_tax}")
    _set(ws, f"S{fee_tax}", "-")
    _set(ws, f"T{fee_tax}", "-")
    _set(ws, f"U{fee_tax}", "-")

    # ── fx rates (same row as 消費税 / 売上) ──
    if ctx.get("fx_rmb") is not None:
        _set(ws, f"V{rate_row}", ctx["fx_rmb"])
    if ctx.get("fx_usd") is not None:
        _set(ws, f"W{rate_row}", ctx["fx_usd"])

    # ── rewrite ALL footer formulas that embed product/fee/rate row numbers ──
    # blank: E15=M29, M27=SUM(M19:N25), M29=SUM(M27:N28), R29=SUM(W19:W21),
    # R30=R28-R29-V22-V31-V32, R31=R30/5*4, R32=R30/5, N31=M28
    _set(ws, "E15", f"=M{total_row}")
    _set(ws, f"M{subtotal_row}", f"=SUM(M{PRODUCT_START}:N{fee_tax})")
    _set(ws, f"M{total_row}", f"=SUM(M{subtotal_row}:N{tax_row})")
    _set(ws, f"R{purchase_sum_row}", f"=SUM(W{PRODUCT_START}:W{last_product})")
    # 利益総額 = 売上(税抜) - 仕入総額 - 運送料V - 輸入手数料V - 関税V
    _set(
        ws,
        f"R{profit_row}",
        f"=R{tax_row}-R{purchase_sum_row}-V{fee_ship}-V{far_profit_row}-V{jp_profit_row}",
    )
    _set(ws, f"R{far_profit_row}", f"=R{profit_row}/5*4")
    _set(ws, f"R{jp_profit_row}", f"=R{profit_row}/5")
    _set(ws, f"N{far_profit_row}", f"=M{tax_row}")

    # optional sales_ex_tax override (売上税抜) — leave template default if absent
    if ctx.get("sales_ex_tax_jpy") is not None:
        _set(ws, f"R{tax_row}", ctx["sales_ex_tax_jpy"])

    out_dir = os.path.dirname(os.path.abspath(out_path)) or "."
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Render 日本請求書 (イノ間) from JSON; not the USD PI path."
    )
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--data", help="JSON context 文件；省略则读 stdin")
    args = ap.parse_args()

    raw = open(args.data, encoding="utf-8").read() if args.data else sys.stdin.read()
    if not raw.strip():
        sys.exit("没有数据（--data 或 stdin 给 JSON）")
    try:
        ctx = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"JSON 解析失败：{e}")

    path = render(ctx, args.template, args.out)
    print(path)


if __name__ == "__main__":
    main()
