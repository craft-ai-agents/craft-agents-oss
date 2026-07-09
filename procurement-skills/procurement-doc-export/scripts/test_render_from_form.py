#!/usr/bin/env python3
"""E2E: form payload → render_from_form → real blank xlsx."""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

SK = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent


class RenderFromFormE2E(unittest.TestCase):
    def _run(self, payload: dict) -> Path:
        from render_from_form import render_payload

        out = Path(tempfile.mkdtemp()) / "out.xlsx"
        path = render_payload(payload, str(out), preview_only=False)
        p = Path(path)
        self.assertTrue(p.exists(), path)
        return p

    def test_quotation_e2e(self):
        blank = SK / "templates/followup/quotation/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        path = self._run(
            {
                "template": "見積書",
                "mode": "render",
                "context": {
                    "date": "2026-07-09",
                    "to": "ACME KK",
                    "currency": "JPY",
                    "items": [
                        {
                            "part": "Q-MPN-1",
                            "qty": 10,
                            "price": 100,
                            "unit": "PCS",
                            "dc": "NA",
                            "lead_time": "2w",
                        }
                    ],
                },
            }
        )
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb.active
        self.assertEqual(ws["C5"].value, "Q-MPN-1")
        self.assertEqual(ws["D5"].value, 10)
        self.assertEqual(ws["H5"].value, 100)
        wb.close()

    def test_far_po_e2e(self):
        blank = SK / "templates/purchase-order/far/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        path = self._run(
            {
                "template": "FAR 采购订单",
                "context": {
                    "po_number": "PO-E2E-1",
                    "po_date": "2026-07-09",
                    "supplier_name": "供方E2E",
                    "items": [{"brand": "ST", "part": "STM32", "qty": 5, "price": 1.2}],
                },
            }
        )
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb.active
        self.assertEqual(ws["D5"].value, "PO-E2E-1")
        self.assertEqual(ws["J6"].value, "供方E2E")
        self.assertEqual(ws["D11"].value, "STM32")
        wb.close()

    def test_domestic_e2e(self):
        blank = SK / "templates/shipping-customs/domestic-delivery/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        path = self._run(
            {
                "template": "国内送货单",
                "context": {
                    "receiver_company": "收货E2E",
                    "receiver_address": "深圳地址",
                    "items": [{"part": "DOM-1", "qty": 3, "unit": "PCS"}],
                },
            }
        )
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb["送货单"]
        self.assertEqual(ws["C3"].value, "收货E2E")
        self.assertEqual(ws["C7"].value, "DOM-1")
        wb.close()

    def test_hedge_e2e(self):
        blank = SK / "templates/INO_SA_应收应付双凯杰对冲结算书模板.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        path = self._run(
            {
                "template": "对冲结算书",
                "context": {
                    "order_no": "INSZ-E2E",
                    "party_b": "乙方E2E",
                    "purchase_lines": [{"part": "P-BUY", "qty": 1, "price": 10}],
                    "sales_lines": [{"part": "P-SELL", "qty": 2, "price": 20}],
                },
            }
        )
        from openpyxl import load_workbook

        wb = load_workbook(path)
        self.assertIn("INSZ-E2E", str(wb["对冲协议书"]["A3"].value))
        self.assertEqual(wb["采购"]["D3"].value, "P-BUY")
        self.assertEqual(wb["销售"]["G2"].value, "P-SELL")
        wb.close()

    def test_preview_only(self):
        from render_from_form import render_payload
        import io
        from contextlib import redirect_stdout

        buf = io.StringIO()
        with redirect_stdout(buf):
            render_payload(
                {
                    "template": "見積書",
                    "context": {"items": [{"part": "X", "qty": 1, "price": 2}]},
                },
                None,
                preview_only=True,
            )
        out = buf.getvalue()
        self.assertIn("見積書", out)
        self.assertIn("X", out)

    def test_unknown_template(self):
        from render_from_form import render_payload

        with self.assertRaises(SystemExit):
            render_payload({"template": "不存在", "context": {}}, "/tmp/x.xlsx", False)


if __name__ == "__main__":
    unittest.main()
