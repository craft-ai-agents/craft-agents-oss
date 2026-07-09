#!/usr/bin/env python3
"""Drive all new render entry points against real blanks; assert fixture fields read back."""
from __future__ import annotations

import json
import subprocess
import tempfile
import unittest
from pathlib import Path

SK = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent


def _run(cmd: list[str], ctx: dict) -> Path:
    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
        json.dump(ctx, f, ensure_ascii=False)
        data_path = f.name
    out = Path(tempfile.mkdtemp()) / "out.xlsx"
    proc = subprocess.run(
        ["uv", "run", "--with", "openpyxl", "python3", *cmd, "--out", str(out), "--data", data_path],
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        raise AssertionError(f"rc={proc.returncode}\n{proc.stdout}\n{proc.stderr}\ncmd={cmd}")
    path = Path(proc.stdout.strip().splitlines()[-1])
    assert path.exists(), path
    return path


def _load(path: Path):
    from openpyxl import load_workbook

    return load_workbook(path)


class FollowupRenders(unittest.TestCase):
    def test_quotation(self):
        blank = SK / "templates/followup/quotation/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        ctx = {
            "date": "2026-07-09",
            "to": "ACME KK",
            "items": [
                {"part": "Q-MPN-1", "qty": 10, "price": 100, "unit": "PCS"},
                {"part": "Q-MPN-2", "qty": 2, "price": 50},
            ],
            "currency": "JPY",
        }
        path = _run(
            [str(SCRIPTS / "render_followup.py"), "--kind", "quotation", "--template", str(blank)],
            ctx,
        )
        wb = _load(path)
        ws = wb.active
        self.assertEqual(ws["C5"].value, "Q-MPN-1")
        self.assertEqual(ws["D5"].value, 10)
        self.assertEqual(ws["H5"].value, 100)
        self.assertEqual(ws["C6"].value, "Q-MPN-2")
        self.assertTrue(str(ws["J7"].value).startswith("=SUM"))
        wb.close()

    def test_takebishi_pi(self):
        blank = SK / "templates/followup/takebishi-pi/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        ctx = {
            "invoice_no": "TB-PI-001",
            "invoice_date": "20260709",
            "po_number": "PO-9",
            "items": [{"part": "TB-PART", "qty": 4, "price": 12.5, "desc": "conn"}],
        }
        path = _run(
            [str(SCRIPTS / "render_followup.py"), "--kind", "takebishi-pi", "--template", str(blank)],
            ctx,
        )
        wb = _load(path)
        ws = wb.active
        self.assertEqual(ws["G7"].value, "TB-PI-001")
        self.assertEqual(ws["C17"].value, "TB-PART")
        self.assertEqual(ws["D17"].value, 4)
        self.assertEqual(ws["G17"].value, 12.5)
        wb.close()

    def test_takebishi_fee(self):
        blank = SK / "templates/followup/takebishi-fee/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        ctx = {
            "invoice_no": "FEE-1",
            "invoice_date": "2026年7月9日",
            "items": [{"part": "手续费-型号A", "qty": 1, "price": 3000, "unit": "式"}],
        }
        path = _run(
            [str(SCRIPTS / "render_followup.py"), "--kind", "takebishi-fee", "--template", str(blank)],
            ctx,
        )
        wb = _load(path)
        ws = wb.active
        self.assertIn("FEE-1", str(ws["K5"].value))
        self.assertEqual(ws["C19"].value, "手续费-型号A")
        self.assertEqual(ws["J19"].value, 3000)
        wb.close()


class PoRenders(unittest.TestCase):
    def test_far(self):
        blank = SK / "templates/purchase-order/far/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        ctx = {
            "po_number": "PO-FAR-99",
            "po_date": "2026-07-09",
            "supplier_name": "测试供应商",
            "items": [
                {"brand": "ST", "part": "STM32", "qty": 100, "price": 3.2, "unit": "PCS"}
            ],
        }
        path = _run(
            [str(SCRIPTS / "render_po.py"), "--kind", "far", "--template", str(blank)],
            ctx,
        )
        wb = _load(path)
        ws = wb.active
        self.assertEqual(ws["D5"].value, "PO-FAR-99")
        self.assertEqual(ws["J6"].value, "测试供应商")
        self.assertEqual(ws["D11"].value, "STM32")
        self.assertEqual(ws["F11"].value, 100)
        self.assertEqual(ws["I11"].value, 3.2)
        wb.close()

    def test_sz_zh_tax_modes(self):
        blank = SK / "templates/purchase-order/sz-ino-contract-zh/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        for mode, expect in (("inclusive", "含税"), ("exclusive", "未税")):
            ctx = {
                "po_number": f"PO-SZ-{mode}",
                "tax_mode": mode,
                "supplier_name": "供方A",
                "items": [{"brand": "X", "part": "P1", "qty": 2, "price": 1.5}],
            }
            path = _run(
                [str(SCRIPTS / "render_po.py"), "--kind", "sz-ino-contract-zh", "--template", str(blank)],
                ctx,
            )
            wb = _load(path)
            ws = wb.active
            self.assertIn("PO-SZ", str(ws["A3"].value))
            self.assertIn(expect, str(ws["A14"].value))
            self.assertEqual(ws["C10"].value, "P1")
            wb.close()

    def test_ino_en_ship_modes(self):
        blank = SK / "templates/purchase-order/ino-contract-en/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        for mode in ("japan", "shenzhen"):
            ctx = {
                "po_number": f"PO-EN-{mode}",
                "ship_to_mode": mode,
                "supplier_name": "Supplier Co",
                "items": [{"part": "EN-PART", "qty": 8, "price": 9.9, "desc": "ic"}],
            }
            path = _run(
                [str(SCRIPTS / "render_po.py"), "--kind", "ino-contract-en", "--template", str(blank)],
                ctx,
            )
            wb = _load(path)
            ws = wb.active
            self.assertEqual(ws["C8"].value, f"PO-EN-{mode}")
            self.assertEqual(ws["B14"].value, "EN-PART")
            self.assertEqual(ws["D14"].value, 8)
            ship = str(ws["C10"].value)
            if mode == "japan":
                self.assertIn("Japan", ship)
            else:
                self.assertIn("Shenzhen", ship)
            wb.close()


class ShippingRenders(unittest.TestCase):
    def test_no_declaration(self):
        blank = SK / "templates/shipping-customs/no-declaration/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        ctx = {
            "sold_to": {"company": "Takebishi Corp", "address": "Kyoto"},
            "po_number": "PO-ND-1",
            "items": [{"part": "ND-PART", "brand": "BrandX", "hs_code": "8542"}],
        }
        path = _run(
            [
                str(SCRIPTS / "render_shipping.py"),
                "--kind",
                "no-declaration",
                "--template",
                str(blank),
            ],
            ctx,
        )
        wb = _load(path)
        ws = wb["汇总"]
        self.assertEqual(ws["C1"].value, "Takebishi Corp")
        self.assertEqual(ws["B20"].value, "PO-ND-1")
        self.assertEqual(ws["B22"].value, "ND-PART")
        wb.close()

    def test_export_currency_marker(self):
        blank = SK / "templates/shipping-customs/export-declaration/blank.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        cases = (
            (
                "cny",
                "出口发票",
                "B16",
                "EX-cny",
                "出口合同",
                "B17",
            ),
            (
                "usd",
                "（美）出口装箱单 ",
                "B13",
                "EX-usd",
                None,
                None,
            ),
            (
                "jpy1039",
                "1039发票",
                "E17",
                "EX-1039",
                "1039发票",
                "D17",
            ),
        )
        for cs, sheet, coord, part, sheet2, coord2 in cases:
            ctx = {
                "currency_set": cs,
                "items": [{"part": part, "hs_code": "8536", "qty": 7, "price": 11}],
            }
            path = _run(
                [
                    str(SCRIPTS / "render_shipping.py"),
                    "--kind",
                    "export-declaration",
                    "--template",
                    str(blank),
                ],
                ctx,
            )
            wb = _load(path)
            self.assertIn("汇总", wb.sheetnames)
            self.assertEqual(wb["汇总"]["A35"].value, f"currency_set={cs}")
            self.assertEqual(wb["汇总"]["B9"].value, part)
            self.assertEqual(wb["汇总"]["B10"].value, "8536")
            # currency-set business sheet must actually change (not 汇总-only)
            self.assertIn(sheet, wb.sheetnames, f"missing pack sheet for {cs}")
            cell_val = str(wb[sheet][coord].value or "")
            self.assertIn(part, cell_val, f"{cs} {sheet}!{coord}={cell_val!r}")
            if sheet2 and coord2:
                if cs == "jpy1039" and coord2 == "D17":
                    self.assertEqual(wb[sheet2][coord2].value, "8536")
                else:
                    self.assertIn(part, str(wb[sheet2][coord2].value or ""))
            # blank row on a non-selected pack sheet should not be our fixture part
            # (smoke: cny must not write EX-cny into USD packing B13 as sole path)
            wb.close()

    def test_import_and_domestic(self):
        for kind, blank_rel, check in (
            (
                "import-declaration",
                "templates/shipping-customs/import-declaration/blank.xlsx",
                lambda wb, ctx: self.assertEqual(
                    wb[[n for n in wb.sheetnames if n.strip() == "汇总"][0]]["B11"].value,
                    "IMP-PART",
                ),
            ),
            (
                "domestic-delivery",
                "templates/shipping-customs/domestic-delivery/blank.xlsx",
                lambda wb, ctx: (
                    self.assertEqual(wb["送货单"]["C3"].value, "收货公司A"),
                    self.assertEqual(wb["送货单"]["C7"].value, "DOM-PART"),
                ),
            ),
        ):
            blank = SK / blank_rel
            if not blank.exists():
                self.skipTest(f"blank missing {blank}")
            if kind == "import-declaration":
                ctx = {
                    "seller": {"company": "DIGIKEY"},
                    "contract_no": "YN-IMP-1",
                    "items": [{"part": "IMP-PART", "qty": 10, "price": 1.2, "hs_code": "8542"}],
                    "part": "IMP-PART",
                    "qty": 10,
                    "price": 1.2,
                    "currency": "USD",
                }
            else:
                ctx = {
                    "receiver_company": "收货公司A",
                    "items": [{"part": "DOM-PART", "qty": 100, "unit": "PCS"}],
                }
            path = _run(
                [str(SCRIPTS / "render_shipping.py"), "--kind", kind, "--template", str(blank)],
                ctx,
            )
            wb = _load(path)
            check(wb, ctx)
            wb.close()


class HedgeRender(unittest.TestCase):
    def test_hedge(self):
        blank = SK / "templates/INO_SA_应收应付双凯杰对冲结算书模板.xlsx"
        if not blank.exists():
            self.skipTest("blank missing")
        ctx = {
            "order_no": "INSZ-TEST-1",
            "order_date": "2026年7月9日",
            "party_b": "测试乙方",
            "contact_b": "张三",
            "purchase_lines": [{"part": "P-BUY", "qty": 1, "price": 10}],
            "sales_lines": [{"part": "P-SELL", "qty": 2, "price": 20}],
        }
        path = _run([str(SCRIPTS / "render_hedge.py"), "--template", str(blank)], ctx)
        wb = _load(path)
        ws = wb["对冲协议书"]
        self.assertIn("INSZ-TEST-1", str(ws["A3"].value))
        self.assertIn("测试乙方", str(ws["G4"].value))
        self.assertEqual(wb["采购"]["D3"].value, "P-BUY")
        self.assertEqual(wb["采购"]["E3"].value, 1)
        self.assertEqual(wb["采购"]["F3"].value, 10)
        self.assertEqual(wb["销售"]["G2"].value, "P-SELL")
        self.assertEqual(wb["销售"]["H2"].value, 2)
        self.assertEqual(wb["销售"]["I2"].value, 20)
        wb.close()


class NoPiHardcode(unittest.TestCase):
    def test_new_scripts_do_not_import_render_pi(self):
        for name in ("render_followup.py", "render_po.py", "render_shipping.py", "render_hedge.py"):
            body = (SCRIPTS / name).read_text(encoding="utf-8")
            after_doc = body.split('"""', 2)[-1]
            self.assertNotIn("import render_pi", after_doc)
            self.assertNotIn("from render_pi", after_doc)
            self.assertNotIn("美金请款发票模板PI", after_doc)


if __name__ == "__main__":
    unittest.main()
