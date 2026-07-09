#!/usr/bin/env python3
"""Tests for shipped render_jp_invoice — drives real blank template + openpyxl."""
from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent
RENDER = SCRIPTS / "render_jp_invoice.py"
BLANK = SKILL_ROOT / "templates/followup/jp-ino-invoice/blank.xlsx"
RENDER_PI = SCRIPTS / "render_pi.py"


def _run_render(ctx: dict, out: Path) -> str:
    raw = json.dumps(ctx, ensure_ascii=False)
    proc = subprocess.run(
        [
            "uv",
            "run",
            "--with",
            "openpyxl",
            "python3",
            str(RENDER),
            "--template",
            str(BLANK),
            "--out",
            str(out),
        ],
        input=raw,
        text=True,
        capture_output=True,
        check=False,
    )
    if proc.returncode != 0:
        raise AssertionError(
            f"render failed rc={proc.returncode}\nstdout={proc.stdout}\nstderr={proc.stderr}"
        )
    path = proc.stdout.strip().splitlines()[-1]
    return path


class RenderJpInvoiceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not BLANK.exists():
            raise unittest.SkipTest(f"blank missing: {BLANK}")
        if not RENDER.exists():
            raise unittest.SkipTest(f"render script missing: {RENDER}")

    def test_blank_sheet_name_is_yangban(self) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(BLANK, read_only=True)
        self.assertIn("样板", wb.sheetnames)
        wb.close()

    def test_one_product_fills_part_qty_and_header(self) -> None:
        from openpyxl import load_workbook

        ctx = {
            "invoice_no": "OD-TEST-001",
            "invoice_date": "2026-07-09",
            "subject": "電子部品の緊急調達",
            "fx_rmb": 22.5,
            "fx_usd": 160.0,
            "items": [
                {
                    "part": "TEST-MPN-AAA",
                    "qty": 7,
                    "unit_price_usd": 12.5,
                }
            ],
            "shipping": {"qty": 0},
            "import_tax_jpy": 0,
        }
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "one.xlsx"
            path = _run_render(ctx, out)
            self.assertTrue(Path(path).exists())
            wb = load_workbook(path, data_only=False)
            ws = wb.active
            self.assertEqual(ws["C19"].value, "TEST-MPN-AAA")
            self.assertEqual(ws["R19"].value, 7)
            self.assertEqual(ws["T19"].value, 12.5)
            self.assertEqual(ws["B19"].value, 1)
            # date / no formatting
            self.assertIn("2026年", str(ws["K4"].value))
            self.assertIn("OD-TEST-001", str(ws["K5"].value))
            # qty/price formulas retained
            self.assertEqual(ws["F19"].value, "=R19")
            self.assertEqual(ws["J19"].value, "=V19")
            self.assertTrue(str(ws["M19"].value).startswith("="))
            wb.close()

    def test_multirow_inserts_and_preserves_all_parts(self) -> None:
        from openpyxl import load_workbook

        parts = [f"PART-{i}" for i in range(1, 6)]  # 5 > default 3 slots
        n = len(parts)
        # expected anchors after insert (PRODUCT_START=19, FEE_COUNT=3, GAP=2)
        fee_ship = 19 + n  # 24
        fee_far = fee_ship + 1
        fee_tax = fee_ship + 2
        footer = 19 + n + 3 + 2  # 29 小計
        rate = footer + 1  # 30
        total = footer + 2  # 31 合計
        profit = footer + 3  # 32
        far_p = footer + 4  # 33
        jp_p = footer + 5  # 34
        ctx = {
            "invoice_no": "OD-MULTI-5",
            "invoice_date": "2026年7月1日",
            "items": [
                {"part": p, "qty": i + 1, "unit_price_rmb": 100 * (i + 1)}
                for i, p in enumerate(parts)
            ],
            "shipping": {"qty": 1, "unit_price_jpy": 500},
            "import_tax_jpy": 0,
        }
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "multi.xlsx"
            path = _run_render(ctx, out)
            wb = load_workbook(path, data_only=False)
            ws = wb.active
            for i, p in enumerate(parts):
                r = 19 + i
                self.assertEqual(ws[f"C{r}"].value, p, msg=f"row {r}")
                self.assertEqual(ws[f"R{r}"].value, i + 1)
                self.assertEqual(ws[f"S{r}"].value, 100 * (i + 1))
                # rate abs refs must track rate_row after insert
                self.assertEqual(
                    ws[f"V{r}"].value,
                    f"=S{r}*$V${rate}+T{r}*$W${rate}+U{r}",
                )
            self.assertEqual(ws[f"C{fee_ship}"].value, "立替運送料")
            self.assertEqual(ws[f"R{fee_ship}"].value, 1)
            self.assertEqual(ws[f"C{fee_far}"].value, "FAR手数料")
            self.assertEqual(ws[f"J{fee_far}"].value, f"=$R${far_p}")
            # 合計 must be 小計+消費税 range — NOT stale blank M27:N28
            self.assertEqual(ws["E15"].value, f"=M{total}")
            self.assertEqual(ws[f"M{footer}"].value, f"=SUM(M19:N{fee_tax})")
            self.assertEqual(ws[f"M{total}"].value, f"=SUM(M{footer}:N{rate})")
            self.assertEqual(ws[f"R{total}"].value, f"=SUM(W19:W{18+n})")
            self.assertEqual(
                ws[f"R{profit}"].value,
                f"=R{rate}-R{total}-V{fee_ship}-V{far_p}-V{jp_p}",
            )
            self.assertEqual(ws[f"R{far_p}"].value, f"=R{profit}/5*4")
            self.assertEqual(ws[f"R{jp_p}"].value, f"=R{profit}/5")
            wb.close()

    def test_render_module_does_not_call_render_pi(self) -> None:
        src = RENDER.read_text(encoding="utf-8")
        # no import / subprocess of PI renderer (docstring may mention contrast)
        self.assertNotIn("import render_pi", src)
        self.assertNotIn("from render_pi", src)
        self.assertNotIn("render_pi.main", src)
        self.assertNotIn("render_pi.py", src.split('"""', 2)[-1])  # body after docstring
        self.assertIn('TPL_SHEET = "样板"', src)
        self.assertNotIn("美金请款发票模板PI", src)

    def test_less_than_default_slots_still_works(self) -> None:
        from openpyxl import load_workbook

        n = 1
        fee_ship = 19 + n  # 20
        fee_far = fee_ship + 1  # 21
        fee_tax = fee_ship + 2  # 22
        footer = 19 + n + 3 + 2  # 25
        rate = footer + 1  # 26
        total = footer + 2  # 27
        profit = footer + 3  # 28
        far_p = footer + 4  # 29
        jp_p = footer + 5  # 30
        ctx = {
            "invoice_no": "OD-ONE-SLOT",
            "invoice_date": "2026-01-02",
            "items": [{"part": "ONLY-ONE", "qty": 3, "unit_price_jpy": 999}],
        }
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "less.xlsx"
            _run_render(ctx, out)
            wb = load_workbook(out, data_only=False)
            ws = wb.active
            self.assertEqual(ws["C19"].value, "ONLY-ONE")
            self.assertEqual(ws["U19"].value, 999)
            self.assertEqual(ws[f"C{fee_ship}"].value, "立替運送料")
            # no circular 合計: blank stale was SUM(M27:N28) which equals total_row when n=1
            self.assertEqual(ws[f"M{total}"].value, f"=SUM(M{footer}:N{rate})")
            self.assertNotIn("M27:N28", str(ws[f"M{total}"].value))
            self.assertEqual(ws["E15"].value, f"=M{total}")
            self.assertEqual(ws[f"M{footer}"].value, f"=SUM(M19:N{fee_tax})")
            self.assertEqual(ws[f"R{total}"].value, "=SUM(W19:W19)")
            self.assertEqual(
                ws[f"R{profit}"].value,
                f"=R{rate}-R{total}-V{fee_ship}-V{far_p}-V{jp_p}",
            )
            self.assertEqual(ws[f"J{fee_far}"].value, f"=$R${far_p}")
            wb.close()


if __name__ == "__main__":
    unittest.main()
