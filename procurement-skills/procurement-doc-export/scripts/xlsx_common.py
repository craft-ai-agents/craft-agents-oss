"""Shared openpyxl helpers for doc-export renders (merge-safe row ops)."""
from __future__ import annotations

import copy
from typing import Any


def set_cell(ws, coord: str, value: Any) -> None:
    """Write value; if coord is inside a merge, write to the merge top-left."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import coordinate_to_tuple, get_column_letter, range_boundaries

    cell = ws[coord]
    if isinstance(cell, MergedCell):
        row, col = coordinate_to_tuple(coord)
        for rng in ws.merged_cells.ranges:
            minc, minr, maxc, maxr = range_boundaries(str(rng))
            if minr <= row <= maxr and minc <= col <= maxc:
                coord = f"{get_column_letter(minc)}{minr}"
                break
    ws[coord] = value


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        if s.has_style:
            d.font = copy.copy(s.font)
            d.fill = copy.copy(s.fill)
            d.border = copy.copy(s.border)
            d.alignment = copy.copy(s.alignment)
            d.number_format = s.number_format
            d.protection = copy.copy(s.protection)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def insert_rows_keep_merges(ws, idx: int, amount: int) -> None:
    from openpyxl.utils import range_boundaries

    moved = []
    for rng in list(ws.merged_cells.ranges):
        minc, minr, maxc, maxr = range_boundaries(str(rng))
        if minr >= idx:
            ws.unmerge_cells(str(rng))
            moved.append((minc, minr, maxc, maxr))
    ws.insert_rows(idx, amount)
    for minc, minr, maxc, maxr in moved:
        ws.merge_cells(
            start_row=minr + amount,
            start_column=minc,
            end_row=maxr + amount,
            end_column=maxc,
        )


def delete_rows_keep_merges(ws, idx: int, amount: int) -> None:
    from openpyxl.utils import range_boundaries

    moved = []
    for rng in list(ws.merged_cells.ranges):
        minc, minr, maxc, maxr = range_boundaries(str(rng))
        if minr >= idx + amount:
            ws.unmerge_cells(str(rng))
            moved.append((minc, minr - amount, maxc, maxr - amount))
        elif minr >= idx:
            ws.unmerge_cells(str(rng))
    ws.delete_rows(idx, amount)
    for minc, minr, maxc, maxr in moved:
        ws.merge_cells(start_row=minr, start_column=minc, end_row=maxr, end_column=maxc)


def adjust_item_block(
    ws,
    *,
    item_start: int,
    default_slots: int,
    n: int,
    total_or_next_row: int,
    max_col: int,
    merge_ranges_for_row=None,
) -> int:
    """Resize item block so it has exactly n rows. Returns new total/next row index."""
    if n < 1:
        raise SystemExit("items 为空：至少要有一个货品行")
    if n > default_slots:
        extra = n - default_slots
        insert_rows_keep_merges(ws, total_or_next_row, extra)
        for i in range(extra):
            dst = item_start + default_slots + i
            copy_row_style(ws, item_start, dst, max_col)
            if merge_ranges_for_row:
                for a, b in merge_ranges_for_row(dst):
                    try:
                        ws.merge_cells(start_row=dst, start_column=a, end_row=dst, end_column=b)
                    except Exception:
                        pass
        return total_or_next_row + extra
    if n < default_slots:
        delete_rows_keep_merges(ws, item_start + n, default_slots - n)
        return total_or_next_row - (default_slots - n)
    return total_or_next_row


def load_single_sheet(template: str, sheet_name: str | None = None):
    from openpyxl import load_workbook
    import os

    if not os.path.exists(template):
        raise SystemExit(f"模板不存在：{template}")
    wb = load_workbook(template)
    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
        keep = ws.title
    else:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"模板缺少 sheet「{sheet_name}」（现有：{wb.sheetnames}）")
        keep = sheet_name
        ws = wb[keep]
    for name in list(wb.sheetnames):
        if name != keep:
            del wb[name]
    return wb, ws


def load_workbook_keep_all(template: str):
    from openpyxl import load_workbook
    import os

    if not os.path.exists(template):
        raise SystemExit(f"模板不存在：{template}")
    return load_workbook(template)
